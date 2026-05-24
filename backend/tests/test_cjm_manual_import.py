from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import sessionmaker

from app.models import (
    Base,
    ClientExpectation,
    CommunicationPoint,
    LPRImportanceFactor,
    LPRProfile,
    Project,
    ProjectBarrier,
    ProjectGoal,
    ProjectKPI,
)
from app.services.imports.cjm_excel_reader import read_cjm_workbook
from app.services.imports.cjm_importer import (
    CJMImporter,
    normalize_external_lpr_aliases,
)
from app.services.imports.cjm_mappings import map_importance_factor_type, map_operational_model
from app.services.imports.cjm_report import build_report_payload
from app.services.imports.cjm_validator import validate_cjm_workbook
from scripts.generate_cjm_mvp_template import generate_template


def _minimal_workbook(path: Path) -> Path:
    generate_template(path)
    workbook = load_workbook(path)
    workbook["01_Паспорт проекта"].append(["project_001", "external_project_001"])
    workbook["02_ЛПР"].append(["lpr_001", "external_lpr_001", "Роль контроля качества"])
    workbook.save(path)
    workbook.close()
    return path


def _append_named_row(worksheet: object, values: dict[str, object]) -> None:
    worksheet.append([values.get(cell.value) for cell in worksheet[1]])


def test_generate_template_creates_excel_file(tmp_path: Path) -> None:
    output_path = tmp_path / "cjm_template.xlsx"

    generated_path = generate_template(output_path)

    assert generated_path == output_path
    assert output_path.exists()
    workbook = load_workbook(output_path)
    assert "00_Инструкция" in workbook.sheetnames
    assert "08_Цели проекта" in workbook.sheetnames
    assert "08_Планы действий" not in workbook.sheetnames
    assert "09_Что нужно дозапросить" in workbook.sheetnames
    assert "04_История опросов" not in workbook.sheetnames
    workbook.close()


def test_reader_reads_test_excel_rows(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "reader.xlsx")

    workbook = read_cjm_workbook(workbook_path)

    assert workbook.sheets["01_Паспорт проекта"].rows[0].values["Project ID"] == "project_001"
    assert workbook.sheets["02_ЛПР"].rows[0].values["LPR ID"] == "lpr_001"


def test_validator_rejects_forbidden_pii_columns(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "pii.xlsx")
    workbook = load_workbook(workbook_path)
    worksheet = workbook["02_ЛПР"]
    worksheet.cell(row=1, column=worksheet.max_column + 1, value="Email")
    workbook.save(workbook_path)
    workbook.close()

    result = validate_cjm_workbook(read_cjm_workbook(workbook_path))

    assert result.has_errors
    assert any(issue.field_name == "Email" for issue in result.issues)


def test_validator_accepts_anonymized_project_and_lpr_codes(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "valid.xlsx")

    result = validate_cjm_workbook(read_cjm_workbook(workbook_path))

    assert not result.has_errors
    assert result.primary_project_id == "project_001"
    assert result.identifiers["lpr_ids"] == {"lpr_001"}


def test_dry_run_does_not_open_database_session(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "dry_run.xlsx")

    def fail_if_called() -> None:
        raise AssertionError("dry-run must not open a DB session")

    result = CJMImporter(
        session_factory=fail_if_called,
        report_dir=tmp_path / "reports",
    ).run(workbook_path, "dry-run")
    repeated_result = CJMImporter(
        session_factory=fail_if_called,
        report_dir=tmp_path / "reports",
    ).run(workbook_path, "dry-run")

    assert result.status == "validated"
    assert result.report_path.exists()
    assert repeated_result.status == "validated"
    assert repeated_result.report_path.exists()


def test_validator_rejects_prohibited_methodology_sheet(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "ai_resume.xlsx")
    workbook = load_workbook(workbook_path)
    workbook.create_sheet("AI-резюме")
    workbook.save(workbook_path)
    workbook.close()

    result = validate_cjm_workbook(read_cjm_workbook(workbook_path))

    assert result.has_errors
    assert any(issue.sheet_name == "AI-резюме" for issue in result.issues)


def test_validator_rejects_action_plan_sheet(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "action_plan.xlsx")
    workbook = load_workbook(workbook_path)
    workbook.create_sheet("08_Планы действий")
    workbook.save(workbook_path)
    workbook.close()

    result = validate_cjm_workbook(read_cjm_workbook(workbook_path))

    assert result.has_errors
    assert any(issue.sheet_name == "08_Планы действий" for issue in result.issues)


def test_validator_accepts_project_goals_sheet(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "goals.xlsx")
    workbook = load_workbook(workbook_path)
    worksheet = workbook["08_Цели проекта"]
    _append_named_row(
        worksheet,
        {
            "Goal ID": "goal_001",
            "Project ID": "project_001",
            "Тип цели": "service",
            "Цель проекта": "Сохранить качество услуги",
        },
    )
    workbook.save(workbook_path)
    workbook.close()

    result = validate_cjm_workbook(read_cjm_workbook(workbook_path))

    assert not result.has_errors
    assert len(result.normalized_sheets["08_Цели проекта"]) == 1


def test_external_lpr_aliases_are_kept_on_one_profile() -> None:
    assert normalize_external_lpr_aliases("845", "55", "845; 55") == "845; 55"


def test_safety_importance_factor_maps_to_code() -> None:
    assert map_importance_factor_type("безопасность") == "safety"


def test_promo_operational_model_maps_to_promo_consulting() -> None:
    assert map_operational_model("промо") == "promo_consulting"


def test_report_describes_unmapped_importance_factor(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "unmapped_importance.xlsx")
    workbook = load_workbook(workbook_path)
    workbook["03_Важности ЛПР"].append(
        ["lpr_001", "external_lpr_001", "нестандартная важность", "Высокая"]
    )
    workbook.save(workbook_path)
    workbook.close()

    validation = validate_cjm_workbook(read_cjm_workbook(workbook_path))
    report = build_report_payload(validation, mode="dry-run", status="validated")

    assert report["unmapped_importance_factors"] == [
        {
            "severity": "warning",
            "sheet_name": "03_Важности ЛПР",
            "row_number": 2,
            "message": "Важность не сопоставлена со справочником, будет загружена как other.",
            "field_name": "Важность",
            "raw_value": "нестандартная важность",
            "current_mapping": "other",
            "issue_type": "unmapped_importance_factor",
            "suggested_action": (
                "Добавить значение в mapping важностей или нормализовать значение в Excel."
            ),
        }
    ]
    assert report["importance_factor_mapping_summary"] == [
        {
            "raw_importance_value": "нестандартная важность",
            "mapped_value": "other",
            "count": 1,
            "example_lpr_id": "lpr_001",
        }
    ]


def test_commit_import_keeps_text_kpi_links(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "text_kpi_links.xlsx")
    workbook = load_workbook(workbook_path)
    workbook["02_ЛПР"].append(["lpr_007", "845", "Региональная роль"])
    workbook["02_ЛПР"].append(["lpr_007", "55", "Региональная роль"])
    workbook["04_Барьеры"].append(
        [
            "barrier_001",
            "Риск задержки",
            "timing",
            "current",
            None,
            "high",
            "lpr_007",
            "845; 55",
            None,
            "scorecard; эффективность мерча",
        ]
    )
    workbook["05_Ожидания клиента"].append(
        [
            "expectation_001",
            "Стабильное качество исполнения",
            "quality",
            "explicit",
            "high",
            "lpr_007",
            "845; 55",
            None,
            "OSA; ISA; PSS",
        ]
    )
    workbook.save(workbook_path)
    workbook.close()

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    test_session_factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)

    result = CJMImporter(
        session_factory=test_session_factory,
        report_dir=tmp_path / "reports",
    ).run(workbook_path, "commit")

    assert result.committed
    with test_session_factory() as session:
        lpr = session.scalar(select(LPRProfile).where(LPRProfile.lpr_code == "lpr_007"))
        lpr_count = session.scalar(
            select(func.count()).select_from(LPRProfile).where(LPRProfile.lpr_code == "lpr_007")
        )
        barrier = session.scalar(
            select(ProjectBarrier).where(ProjectBarrier.source_id == "barrier_001")
        )
        expectation = session.scalar(
            select(ClientExpectation).where(
                ClientExpectation.expectation_text == "Стабильное качество исполнения"
            )
        )

    assert lpr is not None
    assert lpr_count == 1
    assert lpr.external_lpr_id == "845; 55"
    assert barrier is not None
    assert barrier.linked_kpi_text == "scorecard; эффективность мерча"
    assert expectation is not None
    assert expectation.linked_kpi_text == "OSA; ISA; PSS"


def test_importer_persists_v6_context_fields(tmp_path: Path) -> None:
    workbook_path = tmp_path / "v6_context.xlsx"
    generate_template(workbook_path)
    workbook = load_workbook(workbook_path)
    _append_named_row(
        workbook["01_Паспорт проекта"],
        {
            "Project ID": "project_002",
            "External project ID": 1991,
            "Рабочий код проекта": "work_002",
            "Направление проекта": "Электроника",
            "Масштаб проекта": "Федеральный",
            "Основная операционная модель": "промо",
            "Этап жизненного цикла": "Развитие",
            "Статус проекта": "Активный",
            "Дата старта": "2019",
            "Краткое описание проекта": "Anonymized summary",
        },
    )
    _append_named_row(
        workbook["02_ЛПР"],
        {
            "LPR ID": "lpr_001",
            "External LPR ID": "854; 556",
            "Роль / зона влияния": "Региональная роль",
            "Уровень влияния": "high",
            "Статус активности": "active",
            "Предполагаемое отношение к OPEN/услуге": "positive",
            "Основание вывода": "Restored source",
            "Комментарий для ручного уточнения": "Manual note",
        },
    )
    _append_named_row(
        workbook["03_Важности ЛПР"],
        {
            "LPR ID": "lpr_001",
            "External LPR ID": "854; 556",
            "Важность": "безопасность",
            "Критичность": "Высокая",
            "Источник вывода": "Restored source",
            "Доказательство / короткая цитата": "Importance evidence",
            "Период / источник": "2026",
            "Уверенность вывода": "high",
        },
    )
    _append_named_row(
        workbook["04_Барьеры"],
        {
            "Barrier ID": "barrier_001",
            "Название барьера": "Риск исполнения",
            "Тип барьера": "Качество исполнения",
            "Временной статус": "Есть сейчас",
            "Описание барьера": "Barrier description",
            "Критичность": "Высокая",
            "Связанный LPR ID": "lpr_001",
            "External LPR ID": "854; 556",
            "Связанная важность ЛПР": "безопасность",
            "Источник": "Restored source",
            "Доказательство / короткая цитата": "Barrier evidence",
            "Статус барьера": "Открыт",
            "Статус актуальности": "current",
            "Уверенность вывода": "high",
        },
    )
    _append_named_row(
        workbook["05_Ожидания клиента"],
        {
            "Expectation ID": "expectation_001",
            "Ожидание клиента": "Ожидается стабильное качество",
            "Тип ожидания": "Качество",
            "Явное или неявное": "Явное",
            "Критичность": "Высокая",
            "Связанный LPR ID": "lpr_001",
            "External LPR ID": "854; 556",
            "Связанная важность ЛПР": "безопасность",
            "Источник": "Restored source",
            "Доказательство / короткая цитата": "Expectation evidence",
            "Статус актуальности": "current",
            "Уверенность вывода": "high",
        },
    )
    _append_named_row(
        workbook["07_Каналы взаимодействия"],
        {
            "Communication ID": "communication_001",
            "Сторона клиента: LPR ID или роль": "lpr_001",
            "External LPR ID": "854; 556",
            "Сторона OPEN: роль": "Проектная роль",
            "Тема взаимодействия": "Статус проекта",
            "Канал": "Встреча",
            "Частота": "3-5 раз в неделю",
            "Критичность": "Высокая",
            "Источник": "Restored source",
            "Статус актуальности": "current",
            "Комментарий": "Communication note",
        },
    )
    workbook.save(workbook_path)
    workbook.close()

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    test_session_factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    result = CJMImporter(
        session_factory=test_session_factory,
        report_dir=tmp_path / "reports",
    ).run(workbook_path, "commit")

    assert result.committed
    with test_session_factory() as session:
        project = session.scalar(select(Project).where(Project.project_code == "project_002"))
        factor = session.scalar(select(LPRImportanceFactor))
        barrier = session.scalar(select(ProjectBarrier))
        expectation = session.scalar(select(ClientExpectation))
        communication = session.scalar(select(CommunicationPoint))

    assert project is not None
    assert project.external_project_id == "1991"
    assert project.primary_operational_model == "promo_consulting"
    assert project.short_description == "Anonymized summary"
    assert factor is not None
    assert factor.evidence_quote == "Importance evidence"
    assert factor.confidence_level == "high"
    assert barrier is not None
    assert barrier.relevance_status == "current"
    assert barrier.evidence_quote == "Barrier evidence"
    assert barrier.source_text == "Restored source"
    assert expectation is not None
    assert expectation.relevance_status == "current"
    assert expectation.confidence_level == "high"
    assert communication is not None
    assert communication.frequency == "3-5 раз в неделю"


def test_repeated_commit_does_not_duplicate_main_cjm_entities(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "idempotent_commit.xlsx")
    workbook = load_workbook(workbook_path)
    workbook["03_Важности ЛПР"].append(
        ["lpr_001", "external_lpr_001", "безопасность", "Высокая"]
    )
    workbook["04_Барьеры"].append(
        [
            "barrier_001",
            "Риск задержки",
            "Сроки",
            "Есть сейчас",
            None,
            "Высокая",
        ]
    )
    workbook["05_Ожидания клиента"].append(
        [
            "expectation_001",
            "Стабильное качество исполнения",
            "Качество",
            "Явное",
            "Высокая",
        ]
    )
    workbook["06_KPI и критерии успеха"].append(["kpi_001", "Качество исполнения"])
    workbook["07_Каналы взаимодействия"].append(
        ["communication_001", "lpr_001", None, "Роль проекта", "Статус проекта"]
    )
    goals = workbook["08_Цели проекта"]
    _append_named_row(
        goals,
        {
            "Goal ID": "goal_001",
            "Project ID": "project_001",
            "Тип цели": "service",
            "Цель проекта": "Сохранить качество услуги",
        },
    )
    workbook.save(workbook_path)
    workbook.close()

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    test_session_factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    importer = CJMImporter(session_factory=test_session_factory, report_dir=tmp_path / "reports")

    first = importer.run(workbook_path, "commit")
    second = importer.run(workbook_path, "commit")

    assert first.committed
    assert second.committed
    with test_session_factory() as session:
        counts = {
            "lpr_profiles": session.scalar(select(func.count()).select_from(LPRProfile)),
            "importance": session.scalar(
                select(func.count()).select_from(LPRImportanceFactor)
            ),
            "barriers": session.scalar(select(func.count()).select_from(ProjectBarrier)),
            "expectations": session.scalar(
                select(func.count()).select_from(ClientExpectation)
            ),
            "kpis": session.scalar(select(func.count()).select_from(ProjectKPI)),
            "communications": session.scalar(
                select(func.count()).select_from(CommunicationPoint)
            ),
            "goals": session.scalar(select(func.count()).select_from(ProjectGoal)),
        }

    assert counts == {
        "lpr_profiles": 1,
        "importance": 1,
        "barriers": 1,
        "expectations": 1,
        "kpis": 1,
        "communications": 1,
        "goals": 1,
    }


def test_commit_skips_manually_updated_rows_without_force(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "manual_protection.xlsx")
    workbook = load_workbook(workbook_path)
    goals = workbook["08_Цели проекта"]
    _append_named_row(
        goals,
        {
            "Goal ID": "goal_001",
            "Project ID": "project_001",
            "Тип цели": "service",
            "Цель проекта": "Excel goal text",
            "Статус актуальности": "actual",
        },
    )
    workbook.save(workbook_path)
    workbook.close()

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    test_session_factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    importer = CJMImporter(session_factory=test_session_factory, report_dir=tmp_path / "reports")

    first = importer.run(workbook_path, "commit")
    assert first.committed

    with test_session_factory() as session:
        goal = session.scalar(select(ProjectGoal).where(ProjectGoal.source_id == "goal_001"))
        assert goal is not None
        goal.goal_text = "Manual goal text"
        goal.updated_by = "demo_open"
        session.commit()

    second = importer.run(workbook_path, "commit")

    assert second.committed
    assert second.database_counts["project_goals"]["skipped_manual"] == 1
    assert any(issue.issue_type == "manual_update_protection" for issue in second.validation.issues)
    with test_session_factory() as session:
        goal = session.scalar(select(ProjectGoal).where(ProjectGoal.source_id == "goal_001"))
        assert goal is not None
        assert goal.goal_text == "Manual goal text"
        assert goal.updated_by == "demo_open"


def test_force_commit_overwrites_manually_updated_rows(tmp_path: Path) -> None:
    workbook_path = _minimal_workbook(tmp_path / "manual_force.xlsx")
    workbook = load_workbook(workbook_path)
    goals = workbook["08_Цели проекта"]
    _append_named_row(
        goals,
        {
            "Goal ID": "goal_001",
            "Project ID": "project_001",
            "Тип цели": "service",
            "Цель проекта": "Excel goal text",
            "Статус актуальности": "actual",
        },
    )
    workbook.save(workbook_path)
    workbook.close()

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    test_session_factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    importer = CJMImporter(session_factory=test_session_factory, report_dir=tmp_path / "reports")

    first = importer.run(workbook_path, "commit")
    assert first.committed

    with test_session_factory() as session:
        goal = session.scalar(select(ProjectGoal).where(ProjectGoal.source_id == "goal_001"))
        assert goal is not None
        goal.goal_text = "Manual goal text"
        goal.updated_by = "demo_open"
        session.commit()

    forced = CJMImporter(
        session_factory=test_session_factory,
        report_dir=tmp_path / "reports",
        force=True,
    ).run(workbook_path, "commit")

    assert forced.committed
    assert forced.force
    assert "skipped_manual" not in forced.database_counts.get("project_goals", {})
    with test_session_factory() as session:
        goal = session.scalar(select(ProjectGoal).where(ProjectGoal.source_id == "goal_001"))
        assert goal is not None
        assert goal.goal_text == "Excel goal text"
        assert goal.updated_by is None
