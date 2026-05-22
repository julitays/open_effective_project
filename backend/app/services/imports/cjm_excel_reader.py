from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


INSTRUCTION_SHEET_NAME = "00_Инструкция"
TEMPLATE_SHEET_COLUMNS: dict[str, tuple[str, ...]] = {
    "01_Паспорт проекта": (
        "Project ID",
        "Направление проекта",
        "Масштаб проекта",
        "Операционная модель",
        "Дополнительные операционные контуры",
        "Этап жизненного цикла",
        "Статус проекта",
        "Дата старта",
        "Краткое описание проекта",
        "Что нужно дозаполнить",
    ),
    "02_ЛПР": (
        "LPR ID",
        "Роль ЛПР",
        "Уровень влияния",
        "Участвовал в блиц-опросах",
        "Участвовал в операционных опросах",
        "Периоды, где встречается ЛПР",
        "Предполагаемое отношение к OPEN/услуге",
        "Основание вывода",
        "Что нужно дозапросить",
    ),
    "03_Важности ЛПР": (
        "LPR ID",
        "Важность",
        "Критичность",
        "Источник",
        "Доказательство из комментария",
        "Период",
        "Уверенность вывода",
        "Что нужно дозапросить",
    ),
    "04_История опросов": (
        "Survey ID",
        "Project ID",
        "Тип опроса",
        "Год",
        "Период / дата",
        "LPR ID",
        "Вопрос",
        "Оценка",
        "Комментарий клиента",
        "Тема комментария",
        "Тональность",
        "Критичность",
        "Комментарий полезен для CJM",
        "Почему полезен / не полезен",
    ),
    "05_Барьеры": (
        "Barrier ID",
        "Название барьера",
        "Тип барьера",
        "Временной статус",
        "Описание барьера",
        "Критичность",
        "Связанный LPR ID",
        "Связанная важность ЛПР",
        "Связанный KPI",
        "Источник",
        "Доказательство из комментария",
        "Период первого появления",
        "Период последнего появления",
        "Статус",
        "Что нужно дозапросить",
    ),
    "06_Ожидания клиента": (
        "Expectation ID",
        "Ожидание клиента",
        "Тип ожидания",
        "Явное или неявное",
        "Критичность",
        "Связанный LPR ID",
        "Связанная важность ЛПР",
        "Связанный KPI",
        "Источник",
        "Доказательство из комментария",
        "Как можно проверить выполнение без ожидания следующего годового опроса",
        "Что нужно дозапросить",
    ),
    "07_KPI": (
        "KPI ID",
        "Название KPI / критерия успеха",
        "Тип KPI",
        "Это KPI из тендера или текущий KPI",
        "Критичность для клиента",
        "Связанное ожидание клиента",
        "Связанный барьер",
        "Источник данных",
        "Что считается нормой",
        "Что считается тревогой",
        "Что считается критическим отклонением",
        "Что нужно дозапросить",
    ),
    "08_Каналы взаимодействия": (
        "Communication ID",
        "Сторона клиента",
        "Сторона OPEN",
        "Тема взаимодействия",
        "Канал",
        "Частота",
        "Критичность",
        "Есть риск разрыва коммуникации",
        "Основание вывода",
        "Что нужно дозапросить",
    ),
    "09_Планы устранения": (
        "Plan ID",
        "Связанный Barrier ID",
        "Связанное ожидание клиента",
        "Тип действия",
        "Описание действия",
        "Ответственная роль",
        "Срок или период",
        "Статус",
        "Как проверить выполнение без ожидания следующего опроса",
        "Ожидаемый эффект",
        "Что нужно дозапросить",
    ),
    "10_Что нужно дозапросить": (
        "Раздел CJM",
        "Что нужно узнать",
        "У кого запросить",
        "Зачем это нужно",
        "Приоритет",
        "Без этого можно продолжать MVP",
    ),
}


@dataclass(slots=True)
class ExcelRow:
    sheet_name: str
    row_number: int
    values: dict[str, object]


@dataclass(slots=True)
class ExcelSheet:
    name: str
    headers: tuple[str, ...]
    rows: list[ExcelRow]


@dataclass(slots=True)
class CJMWorkbookData:
    file_path: Path
    sheets: dict[str, ExcelSheet]


def _clean_cell(value: object) -> object:
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _headers_from_row(values: tuple[object, ...]) -> tuple[str, ...]:
    headers = [str(value).strip() if value is not None else "" for value in values]
    while headers and not headers[-1]:
        headers.pop()
    return tuple(headers)


def read_cjm_workbook(file_path: str | Path) -> CJMWorkbookData:
    path = Path(file_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, ExcelSheet] = {}

    for worksheet in workbook.worksheets:
        row_iter = worksheet.iter_rows(values_only=True)
        first_row = next(row_iter, ())
        headers = _headers_from_row(first_row)
        rows: list[ExcelRow] = []

        for row_number, values in enumerate(row_iter, start=2):
            cells = tuple(_clean_cell(value) for value in values[: len(headers)])
            if not any(value not in (None, "") for value in cells):
                continue
            row_values = {header: cells[index] for index, header in enumerate(headers) if header}
            rows.append(ExcelRow(worksheet.title, row_number, row_values))

        sheets[worksheet.title] = ExcelSheet(worksheet.title, headers, rows)

    workbook.close()
    return CJMWorkbookData(file_path=path, sheets=sheets)
