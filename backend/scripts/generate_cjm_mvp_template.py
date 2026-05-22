from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.services.imports.cjm_excel_reader import (  # noqa: E402
    INSTRUCTION_SHEET_NAME,
    TEMPLATE_SHEET_COLUMNS,
)

DEFAULT_TEMPLATE_PATH = BACKEND_DIR / "data" / "templates" / "cjm_mvp_collection_template.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True, color="1F2937")
INSTRUCTION_FILL = PatternFill("solid", fgColor="E8F5E9")

VALIDATION_OPTIONS: dict[tuple[str, str], tuple[str, ...]] = {
    ("01_Паспорт проекта", "Направление проекта"): (
        "FMCG",
        "Электроника",
        "КАФ",
        "Промо",
        "Обучение",
        "Аудит",
        "Смешанное",
        "Другое",
    ),
    ("01_Паспорт проекта", "Масштаб проекта"): (
        "Локальный",
        "Региональный",
        "Федеральный",
        "Межрегиональный",
        "Неизвестно",
    ),
    ("01_Паспорт проекта", "Основная операционная модель"): (
        "Мерчендайзинг",
        "Совмещенный мерчендайзинг",
        "Промо / консультирование",
        "КАФ / кадровое администрирование",
        "Обучение",
        "Аудит / контроль качества",
        "Аналитика / отчетность",
        "Смешанная",
        "Другое",
    ),
    ("01_Паспорт проекта", "Этап жизненного цикла"): (
        "Запуск",
        "Стабилизация",
        "Развитие",
        "Удержание",
        "Перезапуск",
        "Риск",
        "Закрытие",
        "Неизвестно",
    ),
    ("01_Паспорт проекта", "Статус проекта"): (
        "Активный",
        "Завершен",
        "Пилот",
        "Под риском",
        "Неизвестно",
    ),
    ("04_Барьеры", "Тип барьера"): (
        "Коммуникация",
        "Качество исполнения",
        "Отчетность",
        "Сроки",
        "Персонал",
        "KPI",
        "Стоимость",
        "Обучение",
        "Контроль",
        "Ожидания",
        "Организация процесса",
        "Другое",
    ),
    ("04_Барьеры", "Временной статус"): (
        "Было",
        "Есть сейчас",
        "Может возникнуть",
        "Повторяется",
    ),
    ("04_Барьеры", "Статус барьера"): (
        "Открыт",
        "В работе",
        "Сдержан",
        "Решен",
        "Мониторинг",
        "Неизвестно",
    ),
    ("05_Ожидания клиента", "Тип ожидания"): (
        "Скорость",
        "Качество",
        "Отчетность",
        "Инициативность",
        "Прозрачность",
        "Стоимость",
        "Экспертиза",
        "Предсказуемость",
        "Минимум ручного контроля",
        "Гибкость",
        "Соблюдение договоренностей",
        "Другое",
    ),
    ("05_Ожидания клиента", "Явное или неявное"): (
        "Явное",
        "Неявное",
        "Неизвестно",
    ),
    ("07_Каналы взаимодействия", "Канал"): (
        "Встреча",
        "Звонок",
        "Почта",
        "Мессенджер",
        "Отчет",
        "BI-дашборд",
        "Другое",
        "Неизвестно",
    ),
    ("07_Каналы взаимодействия", "Частота"): (
        "По запросу",
        "Еженедельно",
        "Раз в две недели",
        "Ежемесячно",
        "Ежеквартально",
        "Редко",
        "Неизвестно",
    ),
    ("08_Планы действий", "Тип действия"): (
        "Устранение",
        "Сдерживание",
        "Профилактика",
        "Эскалация",
        "Коммуникационное действие",
        "Обучение",
        "Контроль",
        "Изменение процесса",
        "Пересмотр KPI",
        "Другое",
    ),
    ("08_Планы действий", "Статус"): (
        "Сделать",
        "В работе",
        "Сделано",
        "Неизвестно",
    ),
    ("08_Планы действий", "Статус подтверждения"): (
        "Подтвержденный план",
        "Черновик",
        "AI-гипотеза",
        "Требует согласования",
    ),
}


def _add_instruction_sheet(workbook: Workbook) -> None:
    worksheet = workbook.active
    worksheet.title = INSTRUCTION_SHEET_NAME
    rows = (
        ("CJM MVP manual intake",),
        ("Заполняйте один обезличенный проект в одном файле.",),
        ("Используйте Project ID вида project_001 и LPR ID вида lpr_001.",),
        ("Не добавляйте ФИО, имена, телефоны, email, названия клиентов и названия проектов.",),
        ("Сначала выполните dry-run; commit разрешен только без критических ошибок.",),
        ("Лист 09 хранит вопросы, которые нужно дозапросить для следующего шага.",),
    )
    for row_index, row in enumerate(rows, start=1):
        worksheet.cell(row=row_index, column=1, value=row[0])
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A1"].fill = INSTRUCTION_FILL
    worksheet.column_dimensions["A"].width = 110
    worksheet.freeze_panes = "A2"


def _style_data_sheet(worksheet: object, columns: tuple[str, ...]) -> None:
    ws = worksheet
    for column_index, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=column_index, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        width = min(max(len(column_name) + 4, 16), 48)
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(columns)).coordinate}"


def _add_list_validation(worksheet: object, column_index: int, options: tuple[str, ...]) -> None:
    escaped = ",".join(options).replace('"', '""')
    validation = DataValidation(type="list", formula1=f'"{escaped}"', allow_blank=True)
    validation.error = "Выберите значение из MVP-справочника."
    validation.prompt = "Можно выбрать значение из списка."
    column_letter = worksheet.cell(row=1, column=column_index).column_letter
    validation.add(f"{column_letter}2:{column_letter}1000")
    worksheet.add_data_validation(validation)


def _add_validations(worksheet: object, sheet_name: str, columns: tuple[str, ...]) -> None:
    for column_index, column_name in enumerate(columns, start=1):
        options = VALIDATION_OPTIONS.get((sheet_name, column_name))
        if column_name in ("Критичность", "Критичность для клиента"):
            options = ("Высокая", "Средняя", "Низкая", "Неизвестно")
        if options is not None:
            _add_list_validation(worksheet, column_index, options)


def generate_template(output_path: str | Path = DEFAULT_TEMPLATE_PATH) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    (BACKEND_DIR / "data" / "imports").mkdir(parents=True, exist_ok=True)
    (BACKEND_DIR / "data" / "import_reports").mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    _add_instruction_sheet(workbook)
    for sheet_name, columns in TEMPLATE_SHEET_COLUMNS.items():
        worksheet = workbook.create_sheet(sheet_name)
        _style_data_sheet(worksheet, columns)
        _add_validations(worksheet, sheet_name, columns)

    workbook.save(path)
    return path


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate the CJM MVP Excel intake template.")
    parser.add_argument("--output", type=Path, default=DEFAULT_TEMPLATE_PATH)
    args = parser.parse_args()
    path = generate_template(args.output)
    print(f"CJM MVP template created: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
